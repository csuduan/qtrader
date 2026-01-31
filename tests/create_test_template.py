import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_test_template():
    output_path = "/opt/dev/qtrader/tests/用户测试案例模板.xlsx"

    # 定义样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    wb = Workbook()

    # ========== Sheet 1: 单元测试用例 ==========
    ws1 = wb.active
    ws1.title = "单元测试用例"

    # 标题
    ws1.merge_cells('A1:J1')
    ws1['A1'] = "单元测试用例模板"
    ws1['A1'].font = Font(bold=True, size=16)
    ws1['A1'].alignment = center_align
    ws1.row_dimensions[1].height = 30

    # 表头
    headers1 = ["模块", "测试类", "测试函数", "测试目的", "前置条件", "输入数据", "预期结果", "实际结果", "状态", "备注"]
    for col, header in enumerate(headers1, 1):
        cell = ws1.cell(row=2, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # 示例数据
    example_data1 = [
        ["test_models", "TestAccountModel", "test_account_creation", "验证账户创建功能", "无", "account_id='test', total_assets=1000.0", "对象创建成功", "", "", ""],
        ["test_models", "TestPositionData", "test_position_creation", "验证持仓数据创建", "无", "symbol='rb2505', volume=10", "对象创建成功", "", "", ""],
        ["test_risk_control", "TestRiskControl", "test_check_order_success", "验证订单风控检查通过", "风控模块初始化", "volume=5, limit=100", "返回True", "", "", ""],
        ["test_risk_control", "TestRiskControl", "test_check_order_exceed_limit", "验证订单数量超限拒绝", "风控模块初始化", "volume=150, limit=100", "返回False", "", "", ""],
        ["test_event_engine", "TestEventEngine", "test_event_registration", "验证事件注册功能", "引擎已启动", "event_type='test', handler=func", "处理器注册成功", "", "", ""],
        ["", "", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", ""],
    ]

    for row_idx, row_data in enumerate(example_data1, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws1.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = left_align if col_idx in [4, 5, 6, 7, 8, 10] else center_align

    # 设置列宽
    col_widths1 = [15, 20, 25, 25, 20, 25, 20, 20, 10, 20]
    for i, width in enumerate(col_widths1, 1):
        ws1.column_dimensions[chr(64+i)].width = width

    # ========== Sheet 2: API测试用例 ==========
    ws2 = wb.create_sheet("API测试用例")

    # 标题
    ws2.merge_cells('A1:K1')
    ws2['A1'] = "API测试用例模板"
    ws2['A1'].font = Font(bold=True, size=16)
    ws2['A1'].alignment = center_align
    ws2.row_dimensions[1].height = 30

    # 表头
    headers2 = ["接口模块", "接口名称", "请求方法", "接口路径", "请求头", "请求参数", "预期响应码", "预期响应体", "实际响应", "状态", "备注"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=2, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # 示例数据
    example_data2 = [
        ["account", "查询账户信息", "GET", "/api/account", "Content-Type: application/json", "无", "200", "{code:0, data:{account_id:'', balance:0}}", "", "", ""],
        ["position", "查询持仓列表", "GET", "/api/positions", "Content-Type: application/json", "无", "200", "{code:0, data:[]}", "", "", ""],
        ["order", "创建订单", "POST", "/api/orders", "Content-Type: application/json", "{symbol:'rb2505', direction:'BUY', volume:1}", "200", "{code:0, data:{order_id:''}}", "", "", ""],
        ["order", "撤单", "DELETE", "/api/orders/{order_id}", "Content-Type: application/json", "无", "200", "{code:0}", "", "", ""],
        ["trade", "查询今日成交", "GET", "/api/trades", "Content-Type: application/json", "无", "200", "{code:0, data:[]}", "", "", ""],
        ["scheduler", "查询任务列表", "GET", "/api/jobs", "Content-Type: application/json", "无", "200", "{code:0, data:[]}", "", "", ""],
        ["scheduler", "暂停任务", "POST", "/api/jobs/{job_id}/pause", "Content-Type: application/json", "无", "200", "{code:0}", "", "", ""],
        ["", "", "", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", "", ""],
    ]

    for row_idx, row_data in enumerate(example_data2, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = left_align if col_idx in [2, 4, 5, 6, 8, 9, 11] else center_align

    # 设置列宽
    col_widths2 = [12, 15, 10, 20, 25, 30, 12, 30, 20, 8, 15]
    for i, width in enumerate(col_widths2, 1):
        ws2.column_dimensions[chr(64+i)].width = width

    # ========== Sheet 3: E2E测试用例 ==========
    ws3 = wb.create_sheet("E2E测试用例")

    # 标题
    ws3.merge_cells('A1:J1')
    ws3['A1'] = "E2E测试用例模板"
    ws3['A1'].font = Font(bold=True, size=16)
    ws3['A1'].alignment = center_align
    ws3.row_dimensions[1].height = 30

    # 表头
    headers3 = ["测试场景", "测试步骤", "操作元素", "输入数据", "预期结果", "实际结果", "状态", "优先级", "截图", "备注"]
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=2, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # 示例数据
    example_data3 = [
        ["用户登录", "1. 打开登录页\n2. 输入用户名\n3. 输入密码\n4. 点击登录", "username输入框\npassword输入框\n登录按钮", "username: admin\npassword: 123456", "登录成功，跳转到首页", "", "", "高", "", ""],
        ["查看账户信息", "1. 登录系统\n2. 点击账户菜单\n3. 查看账户信息", "账户菜单\n账户信息卡片", "无", "显示账户ID、余额、可用资金", "", "", "高", "", ""],
        ["创建订单", "1. 进入下单页面\n2. 选择合约\n3. 输入数量\n4. 点击提交", "合约选择框\n数量输入框\n提交按钮", "symbol: rb2505\nvolume: 1\ndirection: BUY", "订单创建成功，显示订单ID", "", "", "高", "", ""],
        ["撤单操作", "1. 进入订单列表\n2. 选择待撤订单\n3. 点击撤单按钮", "订单列表\n撤单按钮", "order_id: xxx", "撤单成功，订单状态更新", "", "", "中", "", ""],
        ["查看持仓", "1. 点击持仓菜单\n2. 查看持仓列表", "持仓菜单\n持仓列表", "无", "显示当前持仓合约、数量、盈亏", "", "", "高", "", ""],
        ["查看成交记录", "1. 点击成交菜单\n2. 查看今日成交", "成交菜单\n成交列表", "无", "显示今日成交记录", "", "", "中", "", ""],
        ["管理定时任务", "1. 进入任务管理\n2. 暂停/恢复任务", "任务列表\n暂停按钮\n恢复按钮", "job_id: xxx", "任务状态更新", "", "", "低", "", ""],
        ["", "", "", "", "", "", "", "", "", ""],
        ["", "", "", "", "", "", "", "", "", ""],
    ]

    for row_idx, row_data in enumerate(example_data3, 3):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = left_align if col_idx in [2, 3, 4, 5, 6, 10] else center_align

    # 设置列宽
    col_widths3 = [15, 25, 20, 20, 25, 20, 8, 8, 8, 15]
    for i, width in enumerate(col_widths3, 1):
        ws3.column_dimensions[chr(64+i)].width = width

    # ========== Sheet 4: 使用说明 ==========
    ws4 = wb.create_sheet("使用说明")

    # 标题
    ws4.merge_cells('A1:D1')
    ws4['A1'] = "测试用例模板使用说明"
    ws4['A1'].font = Font(bold=True, size=16)
    ws4['A1'].alignment = center_align
    ws4.row_dimensions[1].height = 30

    instructions = [
        ("", ""),
        ("模板结构说明", ""),
        ("本模板包含4个工作表:", ""),
        ("1. 单元测试用例", "用于记录单元测试的详细案例，包含模块、测试类、测试函数等信息"),
        ("2. API测试用例", "用于记录API接口测试案例，包含接口路径、请求方法、参数、预期结果等"),
        ("3. E2E测试用例", "用于记录端到端测试案例，包含测试场景、步骤、操作元素等"),
        ("4. 使用说明", "本说明文档"),
        ("", ""),
        ("填写规范", ""),
        ("【单元测试用例】", ""),
        ("- 模块", "被测试的模块名称，如 test_models, test_risk_control"),
        ("- 测试类", "测试类名称，如 TestAccountModel"),
        ("- 测试函数", "具体的测试函数名，如 test_account_creation"),
        ("- 测试目的", "简要说明该测试的目的"),
        ("- 前置条件", "执行测试前需要满足的条件"),
        ("- 输入数据", "测试使用的输入数据"),
        ("- 预期结果", "期望得到的结果"),
        ("- 实际结果", "实际测试得到的结果（测试后填写）"),
        ("- 状态", "通过/失败/未执行"),
        ("- 备注", "其他补充信息"),
        ("", ""),
        ("【API测试用例】", ""),
        ("- 接口模块", "所属功能模块，如 account, order, trade"),
        ("- 接口名称", "接口的功能描述"),
        ("- 请求方法", "HTTP方法: GET/POST/PUT/DELETE"),
        ("- 接口路径", "API端点路径，如 /api/orders"),
        ("- 请求头", "需要的请求头信息"),
        ("- 请求参数", "请求体或查询参数"),
        ("- 预期响应码", "期望的HTTP状态码"),
        ("- 预期响应体", "期望返回的数据结构"),
        ("- 实际响应", "实际返回的结果（测试后填写）"),
        ("- 状态", "通过/失败/未执行"),
        ("- 备注", "其他补充信息"),
        ("", ""),
        ("【E2E测试用例】", ""),
        ("- 测试场景", "测试的业务场景描述"),
        ("- 测试步骤", "详细的操作步骤，每步一行"),
        ("- 操作元素", "需要操作的页面元素"),
        ("- 输入数据", "需要输入的数据"),
        ("- 预期结果", "期望看到的结果"),
        ("- 实际结果", "实际测试结果（测试后填写）"),
        ("- 状态", "通过/失败/未执行"),
        ("- 优先级", "高/中/低"),
        ("- 截图", "是否需要截图"),
        ("- 备注", "其他补充信息"),
        ("", ""),
        ("测试执行流程", ""),
        ("1. 准备阶段", "- 准备测试环境\n- 准备测试数据\n- 确认测试范围"),
        ("2. 执行阶段", "- 按照测试用例执行测试\n- 记录实际结果\n- 标记测试状态"),
        ("3. 缺陷管理", "- 记录发现的缺陷\n- 跟踪缺陷修复\n- 验证修复结果"),
        ("4. 报告阶段", "- 统计测试通过率\n- 生成测试报告\n- 总结经验教训"),
        ("", ""),
        ("注意事项", ""),
        ("1. 用例设计原则", "- 每个用例只测试一个功能点\n- 用例描述要清晰明确\n- 预期结果要可验证"),
        ("2. 数据管理", "- 敏感信息不要记录在文件中\n- 测试数据要定期更新\n- 重要数据要备份"),
        ("3. 版本控制", "- 定期保存文件版本\n- 重大修改前备份\n- 记录修改历史"),
    ]

    row = 1
    for item in instructions:
        if item[0] == "" and item[1] == "":
            row += 1
            continue

        ws4.cell(row=row, column=1, value=item[0])
        ws4.cell(row=row, column=2, value=item[1])

        # 设置样式
        if item[0] in ["模板结构说明", "填写规范", "测试执行流程", "注意事项"]:
            ws4.cell(row=row, column=1).font = Font(bold=True, size=12, color="4472C4")
        elif item[0].startswith("【"):
            ws4.cell(row=row, column=1).font = Font(bold=True, size=11)
        elif item[0].startswith(("1.", "2.", "3.", "4.")):
            ws4.cell(row=row, column=1).font = Font(bold=True)

        ws4.cell(row=row, column=1).alignment = left_align
        ws4.cell(row=row, column=2).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        row += 1

    # 设置列宽
    ws4.column_dimensions['A'].width = 25
    ws4.column_dimensions['B'].width = 80
    ws4.column_dimensions['C'].width = 15
    ws4.column_dimensions['D'].width = 15

    # 保存文件
    wb.save(output_path)
    print(f"文件已保存到: {output_path}")

if __name__ == "__main__":
    create_test_template()
